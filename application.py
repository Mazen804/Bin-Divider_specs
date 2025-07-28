import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import copy

# Set page configuration
st.set_page_config(page_title="Bin Divider Specification Generator", page_icon=":package:", layout="wide")

# Custom CSS to constrain app width
st.markdown("""
<style>
    .main .block-container {
        max-width: 800px;
        padding-left: 10%;
        padding-right: 10%;
    }
</style>
""", unsafe_allow_html=True)

# Title
st.title("Bin Divider Specification Generator")
st.write("Dynamically add groups and bin box types to generate an Excel file.")

# Initialize session state
if 'groups' not in st.session_state:
    st.session_state.groups = []

# Function to calculate derived fields (for app logic)
def calculate_fields(group_data, bin_data):
    bin_data['# of Aisles'] = group_data['End Aisle'] - group_data['Start Aisle'] + 1
    bin_data['Qty Per Bay'] = bin_data['# of Shelves per Bay'] * bin_data['Qty bins per Shelf']
    bin_data['Total Quantity'] = bin_data['Qty Per Bay'] * group_data['# of Bays']
    bin_data['Bin Gross CBM'] = (bin_data['Depth (mm)'] * bin_data['Height (mm)'] * bin_data['Width (mm)']) / 1_000_000
    bin_data['Bin Net CBM'] = bin_data['Bin Gross CBM'] * bin_data['UT']
    return bin_data

# Function to generate Excel file
def generate_excel(groups):
    columns = [
        'Group Name', 'Floor', 'Mod', 'Depth', 'Start Aisle', 'End Aisle', '# of Aisles', '# of Bays',
        'Total # of Shelves per Bay', 'Bay Design', 'Bin Box Type', 'Depth (mm)',
        'Height (mm)', 'Width (mm)', 'Lip (cm)', '# of Shelves per Bay',
        'Qty bins per Shelf', 'Qty Per Bay', 'Total Quantity', 'UT',
        'Bin Gross CBM', 'Bin Net CBM'
    ]
    df = pd.DataFrame(columns=columns)
    group_row_counts = []
    for group in groups:
        group_data = group['group_data']
        bin_rows = group['bins'] if group['bins'] else [{}]
        for bin_data in bin_rows:
            row = {**group_data, **bin_data}
            row['Lip (cm)'] = '-' if row.get('Lip (cm)', 0.0) == 0.0 else row.get('Lip (cm)', 0.0)
            for col in columns:
                if col not in row:
                    row[col] = None
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        group_row_counts.append(len(bin_rows))

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Bin Box"

    # Write DataFrame to Excel and set formulas
    current_row = 1
    group_start_rows = []
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(r)
        if r_idx > 1:  # Skip header
            group_start_rows.append(current_row if r_idx == 2 else group_start_rows[-1])
            # Set formulas for calculated fields
            ws.cell(row=current_row, column=7).value = f'=F{current_row}-E{current_row}+1'  # # of Aisles
            ws.cell(row=current_row, column=18).value = f'=P{current_row}*Q{current_row}'  # Qty Per Bay
            ws.cell(row=current_row, column=19).value = f'=R{current_row}*$H${group_start_rows[-1]}'  # Total Quantity
            ws.cell(row=current_row, column=21).value = f'=(L{current_row}*M{current_row}*N{current_row})/1000000'  # Bin Gross CBM
            ws.cell(row=current_row, column=22).value = f'=U{current_row}*T{current_row}'  # Bin Net CBM
        current_row += 1

    # Merge and center cells for columns A to I
    current_row = 2
    for row_count in group_row_counts:
        if row_count > 0:
            for col_idx in range(1, 10):  # Columns A to I
                ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + row_count - 1, end_column=col_idx)
                ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
            current_row += row_count

    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Dynamic group and bin input
st.subheader("Manage Groups and Bin Box Types")
if st.button("Add New Group"):
    st.session_state.groups.append({
        'group_data': {
            'Group Name': '',
            'Floor': '',
            'Mod': '',
            'Depth': '',
            'Start Aisle': 1,
            'End Aisle': 1,
            '# of Bays': 1,
            'Total # of Shelves per Bay': 1,
            'Bay Design': ''
        },
        'bins': [],
        'bin_count': 0,
        'finalized': False
    })

# Display and edit groups
for group_idx, group in enumerate(st.session_state.groups):
    # Set expanded=True for newly copied groups (last group if it was just copied)
    is_new_copy = group_idx == len(st.session_state.groups) - 1 and st.session_state.get('last_action') == f"copy_{group_idx-1}"
    with st.expander(f"Group {group_idx + 1}: {group['group_data']['Group Name'] or 'Untitled'} ({'Finalized' if group['finalized'] else 'Editing'})", expanded=not group['finalized'] or is_new_copy):
        if not group['finalized']:
            # Group inputs
            st.write("**Group Details**")
            cols = st.columns(2)
            with cols[0]:
                group['group_data']['Group Name'] = st.text_input("Group Name", value=group['group_data']['Group Name'], key=f"group_name_{group_idx}")
                group['group_data']['Floor'] = st.text_input("Floor", value=group['group_data']['Floor'], key=f"floor_{group_idx}")
                group['group_data']['Mod'] = st.text_input("Mod", value=group['group_data']['Mod'], key=f"mod_{group_idx}")
                group['group_data']['Depth'] = st.text_input("Depth", value=group['group_data']['Depth'], key=f"depth_{group_idx}")
            with cols[1]:
                group['group_data']['Start Aisle'] = st.number_input("Start Aisle", min_value=1, value=int(group['group_data']['Start Aisle']), step=1, key=f"start_aisle_{group_idx}")
                group['group_data']['End Aisle'] = st.number_input("End Aisle", min_value=1, value=int(group['group_data']['End Aisle']), step=1, key=f"end_aisle_{group_idx}")
                group['group_data']['# of Bays'] = st.number_input("# of Bays", min_value=1, value=int(group['group_data']['# of Bays']), step=1, key=f"bays_{group_idx}")
                group['group_data']['Total # of Shelves per Bay'] = st.number_input("Total # of Shelves per Bay", min_value=1, value=int(group['group_data']['Total # of Shelves per Bay']), step=1, key=f"shelves_bay_{group_idx}")
                group['group_data']['Bay Design'] = st.text_input("Bay Design", value=group['group_data']['Bay Design'], key=f"bay_design_{group_idx}")

            # Bin box type count
            group['bin_count'] = st.selectbox("Number of Bin Box Types", options=[0, 1, 2, 3, 4, 5], index=group['bin_count'], key=f"bin_count_{group_idx}")
            
            # Bin box type inputs
            if group['bin_count'] > len(group['bins']):
                group['bins'].extend([{} for _ in range(group['bin_count'] - len(group['bins']))])
            elif group['bin_count'] < len(group['bins']):
                group['bins'] = group['bins'][:group['bin_count']]

            for bin_idx in range(group['bin_count']):
                st.write(f"**Bin Box Type {bin_idx + 1}**")
                cols_bin = st.columns(2)
                bin_data = group['bins'][bin_idx]
                with cols_bin[0]:
                    bin_data['Bin Box Type'] = st.text_input("Bin Box Type", value=bin_data.get('Bin Box Type', ''), placeholder="e.g., 60Deep w/o lip 600*440", key=f"bin_type_{group_idx}_{bin_idx}")
                    bin_data['Depth (mm)'] = st.number_input("Depth (mm)", min_value=0.0, value=bin_data.get('Depth (mm)', 0.0), step=0.1, key=f"depth_mm_{group_idx}_{bin_idx}")
                    bin_data['Height (mm)'] = st.number_input("Height (mm)", min_value=0.0, value=bin_data.get('Height (mm)', 0.0), step=0.1, key=f"height_mm_{group_idx}_{bin_idx}")
                    has_lip = st.checkbox("Has Lip?", value=bin_data.get('Lip (cm)', 0) > 0, key=f"has_lip_{group_idx}_{bin_idx}")
                with cols_bin[1]:
                    bin_data['Width (mm)'] = st.number_input("Width (mm)", min_value=0.0, value=bin_data.get('Width (mm)', 0.0), step=0.1, key=f"width_mm_{group_idx}_{bin_idx}")
                    bin_data['Lip (cm)'] = (bin_data['Height (mm)'] * 0.2 / 10) if has_lip else 0.0
                    st.number_input("Lip (cm)", value=bin_data['Lip (cm)'], disabled=True, key=f"lip_cm_{group_idx}_{bin_idx}")
                    bin_data['# of Shelves per Bay'] = st.number_input("# of Shelves per Bay", min_value=1, value=bin_data.get('# of Shelves per Bay', 1), step=1, key=f"shelves_per_bay_{group_idx}_{bin_idx}")
                    bin_data['Qty bins per Shelf'] = st.number_input("Qty bins per Shelf", min_value=1, value=bin_data.get('Qty bins per Shelf', 1), step=1, key=f"qty_bins_{group_idx}_{bin_idx}")
                    bin_data['UT'] = st.number_input("UT", min_value=0.0, value=bin_data.get('UT', 0.0), step=0.01, key=f"ut_{group_idx}_{bin_idx}")

            if st.button(f"Finalize Group {group_idx + 1}", key=f"finalize_{group_idx}"):
                for bin_idx in range(group['bin_count']):
                    group['bins'][bin_idx] = calculate_fields(group['group_data'], group['bins'][bin_idx])
                group['finalized'] = True
                st.success(f"Group {group_idx + 1} finalized!")
                st.rerun()
        else:
            if st.button(f"Edit Group {group_idx + 1}", key=f"edit_{group_idx}"):
                group['finalized'] = False
                st.success(f"Group {group_idx + 1} opened for editing!")
                st.rerun()

# Display finalized groups
if st.session_state.groups:
    st.subheader("All Groups")
    for i, group in enumerate(st.session_state.groups):
        with st.expander(f"Group {i + 1}: {group['group_data']['Group Name'] or 'Untitled'} ({'Finalized' if group['finalized'] else 'Editing'})"):
            st.write("**Group Details**")
            st.json(group['group_data'])
            if group['bins']:
                st.write("**Bin Box Types**")
                for j, bin_data in enumerate(group['bins']):
                    st.write(f"Bin {j + 1}:")
                    st.json(bin_data)
            if st.button(f"Copy Group {i + 1}", key=f"copy_{i}"):
                new_group = copy.deepcopy(group)
                new_group['finalized'] = False
                new_group['group_data']['Group Name'] = f"{new_group['group_data']['Group Name']} (Copy)" if new_group['group_data']['Group Name'] else "Untitled (Copy)"
                st.session_state.groups.append(new_group)
                st.session_state['last_action'] = f"copy_{i}"  # Track last copy action
                st.success(f"Group {i + 1} copied!")
                st.rerun()

# Download Excel file
if st.session_state.groups:
    excel_data = generate_excel(st.session_state.groups)
    st.download_button(
        label="Download Excel File",
        data=excel_data,
        file_name="Bin_Divider_Specs.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Clear all data
if st.button("Clear All Data"):
    st.session_state.groups = []
    st.rerun()